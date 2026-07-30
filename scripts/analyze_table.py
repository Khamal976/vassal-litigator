#!/usr/bin/env python3
"""
analyze_table.py — аналитическая нога табличного разбора (G.3).

Разделение с `extract_text.py` (см. `shared/ocr.md` §5а): приём таблицу **не
анализирует** — он кладёт в зеркало структурную сводку. Здесь считается то, что
нужно `study-evidence`: арифметика, сверки, флаги. Эвристика чтения (поиск шапки,
нормализация ячеек) **импортируется** из `extract_text.py`, а не дублируется —
иначе приём и анализ разойдутся в понимании того, где у таблицы шапка.

Принцип: скрипт **проверяет**, а не выдумывает. Он не назначает формулу расчёта —
он подбирает ту модель, которая сходится с данными, и честно флагует, если не
сходится ни одна. Все цифры отдаются с адресом клетки (лист + строка + колонка):
для таблицы это то же, что «л. 4» для бумажного документа.

Сети не касается. Ключевая ставка **не сверяется** здесь: скрипт возвращает
использованные в расчёте ставки списком, сверку с cbr.ru делает агент по канону
`shared/conventions.md` → «Расчёт денежных требований: дата начала и ставка (F.10)».

Профили (все четыре — по решению Сюзерена 2026-07-30):
    debt-calc   — расчёт долга / неустойки / процентов (наш и оппонента)
    statement   — выписка по счёту: обороты, полнота, дубли, дробление, получатели
    payments    — платежи против периодов подозрительности гл. III.1 ФЗ-127
    registry    — реестр требований кредиторов: очереди, доля голоса, дубли

Запуск:
    python analyze_table.py <файл.xlsx> --profile debt-calc [--sheet "Лист"]
                            [--tolerance 1.0] [--out-dir DIR]
    python analyze_table.py <файл.xlsx> --profile payments
                            --case-opened 2026-06-15 | --case путь/к/case.yaml
    python analyze_table.py <файл.xlsx> --profile registry --our "Наш клиент"
    python analyze_table.py --selftest

На Windows вызывать `python`, не `python3` (см. `shared/conventions.md` →
«Единый паттерн feature detection + fallback», п. 0).

Выход: JSON. Ненайденное — `null`, не догадка. Профиль `payments` без даты
принятия заявления периоды НЕ считает: от неё зависит применимый состав,
и ошибка на день меняет квалификацию.
"""

import sys
import os
import json
import re
import datetime as dt
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from extract_text import (                      # noqa: E402  (общая эвристика чтения)
    _cell_str, _cell_empty, _find_header_row, _pad_row, _is_date_like,
    TABLE_MAX_SCAN_ROWS, ensure_user_site, dependency_hint,
)

# Подключить user-site, если он есть на диске, но исключён из sys.path (боевой случай
# 2026-07-30: openpyxl установлен, а импорт не проходит). Импорт extract_text это уже
# делает на уровне модуля; вызов здесь — чтобы порядок не зависел от того, кто первый.
ensure_user_site()

# --- параметры ---------------------------------------------------------------

DEFAULT_TOLERANCE = 1.0        # ₽ — абсолютный допуск на строку
REL_TOLERANCE = 0.005          # 0,5 % — относительный допуск
MIN_ROWS_FOR_FORMULA = 2       # меньше — модель не выводим, только проверяем даты/итоги

# --- распознавание колонок --------------------------------------------------

COLUMN_MARKERS = {
    "base":      ("сумма долга", "задолженность", "основной долг", "база", "сумма задолж",
                  "недоимка", "сумма основного", "остаток долга", "долг"),
    "rate":      ("ставка", "%", "процент", "ключевая", "рефинанс"),
    "days":      ("дней", "дни", "кол-во дней", "количество дней", "число дней",
                  "период просрочки", "просрочка, дн"),
    "date_from": ("с ", "дата с", "начало", "период с", "с даты", "дата начала", "от"),
    "date_to":   ("по ", "дата по", "окончание", "период по", "по дату", "дата окончания",
                  "конец"),
    "amount":    ("пени", "неустойка", "проценты", "сумма пени", "начислено",
                  "к взысканию", "итого по строке", "сумма процентов", "размер"),
    "paid":      ("оплачено", "погашено", "платеж", "платёж", "оплата", "внесено"),
}


def _norm_header(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().lower())


def detect_columns(headers: list, rows: list) -> dict:
    """Индексы значимых колонок. Сначала по заголовку, затем добор по типу данных."""
    found = {}
    norm = [_norm_header(h) for h in headers]

    for key, markers in COLUMN_MARKERS.items():
        for idx, head in enumerate(norm):
            if not head or idx in found.values():
                continue
            if any(m in head for m in markers):
                found[key] = idx
                break

    # Добор по данным: даты — колонки, где преобладают date-значения.
    date_cols = []
    for col in range(len(headers)):
        values = [r[col] for r in rows if col < len(r)]
        dates = [v for v in values if _is_date_like(v)]
        if dates and len(dates) >= max(2, len([v for v in values if not _cell_empty(v)]) // 2):
            date_cols.append(col)
    if "date_from" not in found and date_cols:
        found["date_from"] = date_cols[0]
    if "date_to" not in found and len(date_cols) > 1:
        found["date_to"] = date_cols[1]

    return found


def _num(value):
    """Число из ячейки. Суммы, сохранённые как текст, — обычное дело для выгрузок.

    Разделитель определяется, а не угадывается. Прежняя версия вырезала точку как
    символ валюты (точка внутри класса символов литеральна), из-за чего «1500000.75»
    превращалось в 150 000 075, а ставка «8.5» — в 85 %. Дефект был тем опаснее, что
    при единообразно отформатированной колонке все цифры завышались согласованно:
    расхождений не возникало, и юрист получал «машинно точную» сумму.

    Поддержано: «1 200 000,00 руб.» (в т.ч. NBSP и narrow NBSP) · «1500000.75» ·
    «1,234.56» (US) · «1.200.000,00» (EU) · «12 %» · «(500)» как -500 (бухгалтерская
    запись отрицательного) · типографский минус U+2212.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if not s:
        return None

    s = s.replace(" ", " ").replace(" ", " ").replace("−", "-")
    negative_parens = bool(re.fullmatch(r"\(\s*[^)]+\s*\)", s))
    if negative_parens:
        s = s[1:-1].strip()

    s = s.replace("%", "")
    s = re.sub(r"(руб\.?|коп\.?|₽|р\.)", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", "", s)
    s = s.replace("'", "").replace("’", "")

    negative = s.startswith("-")
    s = s.lstrip("+-")
    if not s or not re.fullmatch(r"[\d.,]+", s):
        return None

    has_dot, has_comma = "." in s, "," in s
    if has_dot and has_comma:
        dec = "." if s.rfind(".") > s.rfind(",") else ","
        s = s.replace("," if dec == "." else ".", "").replace(dec, ".")
    elif has_comma or has_dot:
        sep = "," if has_comma else "."
        parts = s.split(sep)
        thousands = len(parts) > 2 and all(len(p) == 3 for p in parts[1:])
        if thousands:
            s = "".join(parts)
        else:
            s = parts[0] + "." + "".join(parts[1:])

    try:
        result = float(s)
    except ValueError:
        return None
    if negative or negative_parens:
        result = -result
    return result


def _as_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        m = re.search(r"(\d{2})[.\-/](\d{2})[.\-/](\d{4})", value)
        if m:
            try:
                return dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            except ValueError:
                return None
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})", value)
        if m:
            try:
                return dt.date(*map(int, m.groups()))
            except ValueError:
                return None
    return None


def _days_in_year(d: dt.date) -> int:
    """Фактическое число дней в году — 365 или 366 (канон F.10, п. 84 ППВС № 7)."""
    year = d.year
    return 366 if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)) else 365


# --- модели формул ----------------------------------------------------------
# Скрипт не навязывает формулу: он проверяет, какая из типовых моделей сходится
# с данными таблицы. Ни одна не сошлась → честный флаг, а не «пересчёт по-своему».

def _formula_models(anchor_date):
    days_in_year = _days_in_year(anchor_date or dt.date(2026, 1, 1))
    return [
        {"id": "annual/actual", "days_in_year": days_in_year,
         "label": f"ставка % годовых / {days_in_year} дн.",
         "fn": lambda b, r, d: b * (r / 100.0) * d / days_in_year},
        {"id": "annual/365", "days_in_year": 365,
         "label": "ставка % годовых / 365 дн.",
         "fn": lambda b, r, d: b * (r / 100.0) * d / 365.0},
        {"id": "annual/366", "days_in_year": 366,
         "label": "ставка % годовых / 366 дн.",
         "fn": lambda b, r, d: b * (r / 100.0) * d / 366.0},
        {"id": "annual/300", "days_in_year": 300,
         "label": "1/300 ставки за день (ЖКХ, налоги)",
         "fn": lambda b, r, d: b * (r / 100.0) * d / 300.0},
        {"id": "daily-percent", "days_in_year": None,
         "label": "ставка % за день (договорная, напр. 0,1 %)",
         "fn": lambda b, r, d: b * (r / 100.0) * d},
        {"id": "daily-fraction", "days_in_year": None,
         "label": "ставка долей за день (напр. 0,001)",
         "fn": lambda b, r, d: b * r * d},
    ]


_REL_TOLERANCE_ACTIVE = REL_TOLERANCE


def _close(a, b, tolerance):
    """Сходятся ли суммы. Относительный допуск отключается явным --tolerance.

    Прежде стояло max(абсолютный, относительный) — из-за чего уменьшить допуск ключом
    было невозможно, и на реестре в 220 млн ₽ проглатывалось расхождение в миллион
    (0,5 % = 1,1 млн). Теперь: --tolerance задан → работает только он.
    """
    if a is None or b is None:
        return False
    limit = tolerance if _REL_TOLERANCE_ACTIVE is None else max(
        tolerance, abs(b) * _REL_TOLERANCE_ACTIVE)
    return abs(a - b) <= limit




TOTAL_ROW_MARKERS = ("итого", "всего", "итог", "оборот за период", "к оплате всего")


def _is_total_row(row: list) -> bool:
    """Строка-итог (в т.ч. промежуточный «Итого за I квартал»).

    Профили обязаны исключать такие строки из данных: иначе итог входит в сумму
    операций, оборот удваивается, и сверка «итог против суммы» даёт расхождение,
    ровно равное верному итогу (поймано состязательной проверкой 2026-07-30).
    """
    joined = " ".join(_cell_str(c).lower() for c in row if not _cell_empty(c))
    return bool(joined) and any(m in joined for m in TOTAL_ROW_MARKERS)


def _find_total_rows(rows: list, header_row1: int, amount_col: int):
    """Все строки-итоги с их значениями → (список, итоговая_строка_или_None).

    Общим итогом считается ПОСЛЕДНЯЯ такая строка: промежуточные («Итого за
    I квартал») стоят выше. Если их несколько, профиль обязан сказать об этом —
    сравнивать сумму всех строк с первым попавшимся «итого» нельзя.
    """
    found = []
    for i, row in enumerate(rows):
        if not _is_total_row(row):
            continue
        val = _num(row[amount_col]) if amount_col is not None and amount_col < len(row) else None
        found.append({"row": header_row1 + 1 + i, "value": val, "index": i})
    grand = None
    for item in reversed(found):
        if item["value"] is not None:
            grand = item
            break
    return found, grand


def _plural(n: int, one: str, few: str, many: str) -> str:
    """«1 платёж / 2 платежа / 5 платежей» — текст читает юрист, падежи заметны."""
    n = abs(int(n))
    if n % 10 == 1 and n % 100 != 11:
        return f"{n} {one}"
    if 2 <= n % 10 <= 4 and not 12 <= n % 100 <= 14:
        return f"{n} {few}"
    return f"{n} {many}"


def _payments_word(n: int) -> str:
    return _plural(n, "платёж", "платежа", "платежей")


def _money(value) -> str:
    """Сумма по-русски: пробел между разрядами, запятая в дробной части.

    Отдельной функцией, потому что формат `{:,.2f}` даёт «13,589.04», а замена
    запятых по всему сообщению съедает и обычные запятые текста (поймано тестом).
    """
    if value is None:
        return "—"
    s = f"{abs(value):,.2f}".replace(",", " ").replace(".", ",")
    return ("−" if value < 0 else "") + s + " ₽"


def _money_signed(value) -> str:
    if value is None:
        return "—"
    return ("+" if value > 0 else "") + _money(value)


# --- адресация --------------------------------------------------------------

def _col_letter(idx0: int) -> str:
    letters, n = "", idx0 + 1
    while n:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _addr(sheet: str, row1: int, col0: int) -> str:
    return f"{sheet}!{_col_letter(col0)}{row1}"


# --- профиль debt-calc ------------------------------------------------------

def profile_debt_calc(sheet_name, headers, rows, header_row1, tolerance) -> dict:
    findings, out_rows = [], []
    cols = detect_columns(headers, rows)

    def add(code, severity, message, cells=None, numbers=None):
        findings.append({"code": code, "severity": severity, "message": message,
                         "cells": cells or [], "numbers": numbers or {}})

    missing = [k for k in ("base", "rate", "days", "amount") if k not in cols]
    if missing:
        add("columns-missing", "warn",
            "Не распознаны колонки: " + ", ".join(missing)
            + ". Проверены заголовки: " + " | ".join(h for h in headers if h)
            + ". Арифметика строк не проверялась — укажите колонки вручную или "
              "приведите шапку к обычным названиям (сумма долга / ставка / дней / пени).")

    # собираем строки данных
    data, skipped_total_rows = [], 0
    for i, row in enumerate(rows):
        if i <= 0 and not row:
            continue
        row1 = header_row1 + 1 + i          # номер строки в файле
        rec = {"row": row1}
        for key, idx in cols.items():
            raw = row[idx] if idx < len(row) else None
            rec[key] = raw
            rec[key + "_num"] = _num(raw)
            if key in ("date_from", "date_to"):
                rec[key + "_date"] = _as_date(raw)
        # строка считается расчётной, если есть база и итог И это не строка-итог:
        # иначе «ИТОГО | 1 000 000 | | | 30 000» войдёт в сумму строк дважды
        if (rec.get("base_num") is not None and rec.get("amount_num") is not None
                and not _is_total_row(row)):
            data.append(rec)
        if _is_total_row(row):
            skipped_total_rows += 1

    # --- выбор модели формулы
    anchor = next((r.get("date_from_date") for r in data if r.get("date_from_date")), None)
    formula = None
    if len(data) >= MIN_ROWS_FOR_FORMULA and not missing:
        scores = []
        for model in _formula_models(anchor):
            ok = 0
            for r in data:
                exp = model["fn"](r["base_num"], r["rate_num"] or 0, r["days_num"] or 0)
                if _close(exp, r["amount_num"], tolerance):
                    ok += 1
            scores.append((ok, model))
        scores.sort(key=lambda s: s[0], reverse=True)
        best_ok, best = scores[0]
        if best_ok >= max(1, int(len(data) * 0.6)):
            formula = {"id": best["id"], "label": best["label"],
                       "matched_rows": best_ok, "total_rows": len(data)}
        else:
            add("formula-unknown", "warn",
                "Ни одна типовая модель расчёта не сошлась с данными "
                f"(лучшая — «{best['label']}»: {best_ok} из {len(data)} строк). "
                "Формула нестандартная либо в таблице ошибки не в одной строке, а в подходе. "
                "Пересчёт по своей формуле НЕ выполнялся — проверьте расчёт вручную.")

    # --- конвенция счёта дней: определяем преобладающую по таблице
    # Обе конвенции законны (включая первый день / без него), поэтому ошибкой
    # является не значение само по себе, а ВЫПАДЕНИЕ строки из конвенции таблицы:
    # именно так выглядит подкрученный на день период у оппонента.
    day_conv = {"inclusive": 0, "exclusive": 0, "other": 0}
    for r in data:
        if r.get("date_from_date") and r.get("date_to_date") and r.get("days_num") is not None:
            incl = (r["date_to_date"] - r["date_from_date"]).days + 1
            stated = int(r["days_num"])
            if stated == incl:
                day_conv["inclusive"] += 1
            elif stated == incl - 1:
                day_conv["exclusive"] += 1
            else:
                day_conv["other"] += 1
    convention = None
    if day_conv["inclusive"] or day_conv["exclusive"]:
        convention = "inclusive" if day_conv["inclusive"] >= day_conv["exclusive"] else "exclusive"

    # --- проверка строк
    for r in data:
        row_findings = []
        d_from, d_to = r.get("date_from_date"), r.get("date_to_date")

        # дни по датам vs указанные дни — против конвенции таблицы
        if d_from and d_to and r.get("days_num") is not None:
            incl = (d_to - d_from).days + 1
            stated = int(r["days_num"])
            expected_days = incl if convention == "inclusive" else incl - 1
            if convention and stated != expected_days:
                row_findings.append({
                    "code": "days-mismatch", "severity": "error",
                    "message": f"строка {r['row']}: указано {stated} дн., а по датам "
                               f"{d_from:%d.%m.%Y}–{d_to:%d.%m.%Y} должно быть "
                               f"{expected_days} дн. В остальных строках таблицы период "
                               f"считается «{'включая первый день' if convention == 'inclusive' else 'со следующего дня'}»"
                               f" — эта строка из конвенции выпадает",
                    "cells": [_addr(sheet_name, r["row"], cols["days"])],
                    "numbers": {"stated": stated, "expected": expected_days,
                                "by_dates_inclusive": incl}})
            elif not convention:
                row_findings.append({
                    "code": "days-unverifiable", "severity": "warn",
                    "message": f"строка {r['row']}: указано {stated} дн., по датам "
                               f"{d_from:%d.%m.%Y}–{d_to:%d.%m.%Y} — {incl} дн. "
                               f"включая первый день ({incl - 1} без него). Ни одна "
                               f"конвенция счёта по таблице не подтверждается — проверьте вручную",
                    "cells": [_addr(sheet_name, r["row"], cols["days"])],
                    "numbers": {"stated": stated, "by_dates_inclusive": incl}})

        # арифметика строки по выбранной модели
        if formula and r.get("rate_num") is not None and r.get("days_num") is not None:
            model = next(m for m in _formula_models(anchor) if m["id"] == formula["id"])
            expected = model["fn"](r["base_num"], r["rate_num"], r["days_num"])
            if not _close(expected, r["amount_num"], tolerance):
                delta = r["amount_num"] - expected
                row_findings.append({
                    "code": "row-arithmetic", "severity": "error",
                    "message": f"строка {r['row']}: в таблице {_money(r['amount_num'])}, "
                               f"по формуле «{formula['label']}» — {_money(expected)}, "
                               f"расхождение {_money_signed(delta)}",
                    "cells": [_addr(sheet_name, r["row"], cols["amount"])],
                    "numbers": {"stated": r["amount_num"], "expected": round(expected, 2),
                                "delta": round(delta, 2)}})

        findings.extend(row_findings)
        out_rows.append({"row": r["row"],
                         "base": r.get("base_num"), "rate": r.get("rate_num"),
                         "days": r.get("days_num"), "amount": r.get("amount_num"),
                         "date_from": d_from.isoformat() if d_from else None,
                         "date_to": d_to.isoformat() if d_to else None,
                         "ok": not row_findings})

    # --- перекрытие и разрывы периодов (двойной счёт дней)
    periods = {"checked": False, "overlaps": [], "gaps": []}
    dated = [r for r in data if r.get("date_from_date") and r.get("date_to_date")]
    if len(dated) >= 2:
        periods["checked"] = True
        dated.sort(key=lambda r: r["date_from_date"])
        for prev, cur in zip(dated, dated[1:]):
            if cur["date_from_date"] <= prev["date_to_date"]:
                overlap = (prev["date_to_date"] - cur["date_from_date"]).days + 1
                periods["overlaps"].append({"rows": [prev["row"], cur["row"]],
                                            "days": overlap})
                add("period-overlap", "error",
                    f"периоды строк {prev['row']} и {cur['row']} перекрываются на "
                    f"{overlap} дн. ({cur['date_from_date']:%d.%m.%Y} ≤ "
                    f"{prev['date_to_date']:%d.%m.%Y}) — двойной счёт дней",
                    [_addr(sheet_name, prev["row"], cols.get("date_to", 0)),
                     _addr(sheet_name, cur["row"], cols.get("date_from", 0))],
                    {"overlap_days": overlap})
            elif (cur["date_from_date"] - prev["date_to_date"]).days > 1:
                gap = (cur["date_from_date"] - prev["date_to_date"]).days - 1
                periods["gaps"].append({"rows": [prev["row"], cur["row"]], "days": gap})
                add("period-gap", "info",
                    f"между строками {prev['row']} и {cur['row']} пропуск {gap} дн. "
                    f"— проверьте, намеренный ли (мораторий, отсрочка) или потерян период",
                    [_addr(sheet_name, cur["row"], cols.get("date_from", 0))],
                    {"gap_days": gap})

    # --- итог: сумма строк vs заявленный итог
    totals = {"rows_sum": None, "stated_total": None, "stated_total_row": None,
              "delta": None}
    if data and "amount" in cols:
        rows_sum = sum(r["amount_num"] for r in data)
        totals["rows_sum"] = round(rows_sum, 2)
        # итог — ПОСЛЕДНЯЯ строка-итог; промежуточные стоят выше
        all_totals, grand = _find_total_rows(rows, header_row1, cols["amount"])
        if len(all_totals) > 1:
            add("intermediate-totals", "info",
                f"строк с пометкой «итого/всего» найдено {len(all_totals)} "
                f"(строки {', '.join(str(t['row']) for t in all_totals)}). За общий итог "
                f"взята последняя; если это не так, укажите нужную строку",
                [], {"rows": [t["row"] for t in all_totals]})
        if grand:
            totals["stated_total"] = grand["value"]
            totals["stated_total_row"] = grand["row"]
        if totals["stated_total"] is not None:
            delta = totals["stated_total"] - rows_sum
            totals["delta"] = round(delta, 2)
            if not _close(totals["stated_total"], rows_sum, tolerance):
                add("total-mismatch", "error",
                    f"итог в таблице {_money(totals['stated_total'])}, сумма строк "
                    f"{_money(rows_sum)}, расхождение {_money_signed(delta)}",
                    [_addr(sheet_name, totals["stated_total_row"], cols["amount"])],
                    {"stated": totals["stated_total"], "rows_sum": round(rows_sum, 2)})

    # --- ставки: перечень для сверки агентом (сеть не трогаем — канон F.10)
    rates = []
    for r in data:
        if r.get("rate_num") is None:
            continue
        rates.append({"row": r["row"], "rate": r["rate_num"],
                      "date_from": r["date_from_date"].isoformat() if r.get("date_from_date") else None,
                      "date_to": r["date_to_date"].isoformat() if r.get("date_to_date") else None})
    distinct = sorted({x["rate"] for x in rates})
    if rates:
        add("rates-to-verify", "info",
            f"в расчёте использованы ставки: {', '.join(f'{v:g} %' for v in distinct)}. "
            "Сверьте значение и дату начала действия каждой по канону F.10 "
            "(база cbr.ru); скрипт сеть не запрашивает.",
            [], {"distinct_rates": distinct})

    # --- частичные оплаты: очерёдность ст. 319 ГК — правовая проверка, не арифметика
    if "paid" in cols and any(r.get("paid_num") for r in data):
        add("partial-payments", "warn",
            "В таблице есть колонка оплат. Проверьте разноску по очерёдности ст. 319 ГК "
            "(канон F.10, п. 7): платёж гасит сначала проценты за пользование, затем "
            "основной долг, и только потом проценты по ст. 395 и неустойку. Если оппонент "
            "закрыл платежами сначала неустойку — остаток долга у него завышен.",
            [_addr(sheet_name, data[0]["row"], cols["paid"])])

    return {"columns": {k: _col_letter(v) for k, v in cols.items()},
            "formula": formula, "rows": out_rows, "findings": findings,
            "totals": totals, "rates": rates, "periods": periods}


# --- чтение листа -----------------------------------------------------------

def _anchor_from_case(case_path: str):
    """Дата принятия заявления из `case.yaml` → (дата, источник).

    Берём именно `case.bankruptcy.case_opened_date`. В схеме у поля стоит
    предупреждение: при нескольких заявлениях о банкротстве там должна лежать дата
    принятия ПЕРВОГО (п. 7 ППВАС № 35). Проверить это машинно нельзя — поэтому
    источник возвращается наружу, чтобы скилл показал его Сюзерену на сверку.
    """
    try:
        import yaml
    except ImportError:
        return None, None, ("PyYAML не установлен — прочитать case.yaml нечем; "
                            "передайте дату ключом --case-opened")
    try:
        with open(case_path, encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
    except Exception as exc:
        return None, None, f"case.yaml не прочитан: {exc}"
    node = (data.get("case") or {}).get("bankruptcy") or {}
    if "case_opened_date" not in node:
        return None, None, ("в case.yaml нет поля case.bankruptcy.case_opened_date "
                            "(блок bankruptcy заполняется только для банкротных дел)")
    value = node.get("case_opened_date")
    if not value:
        return None, None, ("поле case.bankruptcy.case_opened_date пусто — дату принятия "
                            "заявления надо внести в карточку дела")
    date = _as_date(value)
    if date is None:
        return None, None, f"значение case_opened_date «{value}» не распознано как дата"
    return date, f"case.yaml → case.bankruptcy.case_opened_date = {value}", None


def check_stale_formulas(path: str, sheet: str = None):
    """Формулы без сохранённых значений — та же ловушка, что проверяет приём.

    Аналитическая нога обязана проверять её самостоятельно: юрист может запустить
    разбор по файлу, который через приём не проходил. Без этого расчёт из 1С, где
    колонка суммы — формулы без кэша, даёт `findings: []` и `errors: 0` (все строки
    молча отброшены), то есть «сошлось» на нечитаемом файле.
    """
    try:
        from extract_text import _count_stale_formulas
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True, keep_links=False)
        names = [sheet] if sheet else wb.sheetnames
        try:
            wb.close()
        except Exception:
            pass
        return _count_stale_formulas(path, names)
    except Exception:
        return {"formulas": 0, "stale": 0, "checked": False, "examples": []}


def _load_csv(path: str):
    """CSV читается ногой напрямую: приём обещает «итоги по массиву — через
    analyze_table.py», а ветки csv здесь не было вовсе (openpyxl падал «not a zip»)."""
    import csv as _csv
    raw = None
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            with open(path, "r", encoding=encoding, newline="") as f:
                raw = f.read()
            break
        except UnicodeDecodeError:
            continue
    if raw is None:
        raise ValueError("не удалось прочитать CSV ни в UTF-8, ни в cp1251")
    try:
        delimiter = _csv.Sniffer().sniff(raw[:4096], delimiters=";,	|").delimiter
    except Exception:
        delimiter = ";" if raw[:4096].count(";") > raw[:4096].count(",") else ","
    rows = [r for r in _csv.reader(raw.splitlines(), delimiter=delimiter)]
    while rows and all(not str(c).strip() for c in rows[-1]):
        rows.pop()
    return rows


def load_sheet(path: str, sheet: str = None):
    if str(path).lower().endswith(".csv"):
        rows = _load_csv(path)
        if not rows:
            return Path(path).name, [], [], 1
        width = max(len(r) for r in rows)
        hdr = _find_header_row(rows)
        headers = _pad_row([_cell_str(c) for c in rows[hdr]], width)
        data_rows = [_pad_row(r, width) for r in rows[hdr + 1:]]
        return Path(path).name, headers, data_rows, hdr + 1
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    try:
        all_sheets = list(wb.sheetnames)
        ws = wb[sheet] if sheet else wb[all_sheets[0]]
        name = ws.title
        load_sheet.last_sheets = all_sheets      # для предупреждения в main()
        rows = []
        for n, row in enumerate(ws.iter_rows(values_only=True)):
            if n >= TABLE_MAX_SCAN_ROWS:
                break
            rows.append(list(row))
    finally:
        try:
            wb.close()
        except Exception:
            pass
    while rows and all(_cell_empty(c) for c in rows[-1]):
        rows.pop()
    if not rows:
        return name, [], [], 1
    width = max(len(r) for r in rows)
    hdr = _find_header_row(rows)
    headers = _pad_row([_cell_str(c) for c in rows[hdr]], width)
    data_rows = [_pad_row(r, width) for r in rows[hdr + 1:]]
    return name, headers, data_rows, hdr + 1


# --- CLI --------------------------------------------------------------------

# --- профиль statement (выписка по счёту) -----------------------------------

STATEMENT_MARKERS = {
    "date":    ("дата", "дата операции", "дата проводки", "дата документа"),
    "debit":   ("списание", "расход", "дебет", "по дебету", "уменьшение"),
    "credit":  ("приход", "поступление", "кредит", "по кредиту", "зачисление",
                "увеличение"),
    "amount":  ("сумма", "сумма операции", "сумма по операции"),
    "party":   ("контрагент", "плательщик", "получатель", "наименование",
                "корреспондент", "бенефициар"),
    "inn":     ("инн",),
    "purpose": ("назначение", "назначение платежа", "основание", "содержание операции"),
    "balance": ("остаток", "сальдо", "исходящий остаток", "баланс"),
}

SPLIT_MIN_PAYMENTS = 3      # сколько платежей одному лицу подряд считать дроблением
SPLIT_WINDOW_DAYS = 2       # в пределах скольких дней
TOP_PARTIES = 10


def _detect_by_markers(headers: list, markers: dict) -> dict:
    found, norm = {}, [_norm_header(h) for h in headers]
    for key, marks in markers.items():
        for idx, head in enumerate(norm):
            if not head or idx in found.values():
                continue
            if any(m in head for m in marks):
                found[key] = idx
                break
    return found


def _extract_party(rec: dict) -> str:
    """Контрагент: своя колонка, иначе — из назначения платежа (кавычки / ИНН)."""
    if rec.get("party") and _cell_str(rec["party"]):
        return _cell_str(rec["party"])
    purpose = _cell_str(rec.get("purpose"))
    if not purpose:
        return ""
    m = re.search(r"[«\"']([^»\"']{3,60})[»\"']", purpose)
    if m:
        return m.group(1).strip()
    m = re.search(r"\b(\d{10}|\d{12})\b", purpose)
    if m:
        return f"ИНН {m.group(1)}"
    return ""


def profile_statement(sheet_name, headers, rows, header_row1, tolerance) -> dict:
    findings = []
    cols = _detect_by_markers(headers, STATEMENT_MARKERS)

    def add(code, severity, message, cells=None, numbers=None):
        findings.append({"code": code, "severity": severity, "message": message,
                         "cells": cells or [], "numbers": numbers or {}})

    # дата: добор по данным, если по заголовку не нашлась
    if "date" not in cols:
        for col in range(len(headers)):
            values = [r[col] for r in rows if col < len(r)]
            if sum(1 for v in values if _is_date_like(v)) >= max(2, len(values) // 3):
                cols["date"] = col
                break

    if "date" not in cols:
        add("columns-missing", "warn",
            "Колонка даты не распознана — обороты по периодам и разрывы не проверялись. "
            "Заголовки: " + " | ".join(h for h in headers if h))

    has_split_columns = "debit" in cols and "credit" in cols
    if not has_split_columns and "amount" not in cols:
        add("columns-missing", "warn",
            "Не найдены ни пара «приход/списание», ни единая колонка суммы — обороты "
            "не посчитаны. Заголовки: " + " | ".join(h for h in headers if h))

    ops = []
    for i, row in enumerate(rows):
        row1 = header_row1 + 1 + i
        rec = {key: (row[idx] if idx < len(row) else None) for key, idx in cols.items()}
        date = _as_date(rec.get("date"))
        debit = _num(rec.get("debit")) if has_split_columns else None
        credit = _num(rec.get("credit")) if has_split_columns else None
        single = _num(rec.get("amount")) if not has_split_columns else None
        if single is not None:
            # единая колонка со знаком: минус — списание
            debit, credit = (abs(single), None) if single < 0 else (None, single)
        if date is None and debit is None and credit is None:
            continue
        if _is_total_row(row):        # «ИТОГО» — не операция: иначе оборот удваивается
            continue
        ops.append({"row": row1, "date": date, "debit": debit, "credit": credit,
                    "party": _extract_party(rec), "purpose": _cell_str(rec.get("purpose")),
                    "balance": _num(rec.get("balance"))})

    dated = [o for o in ops if o["date"]]
    turnover = {
        "operations": len(ops),
        "credit_total": round(sum(o["credit"] or 0 for o in ops), 2),
        "debit_total": round(sum(o["debit"] or 0 for o in ops), 2),
        "period_from": min((o["date"] for o in dated), default=None),
        "period_to": max((o["date"] for o in dated), default=None),
    }
    turnover["period_from"] = turnover["period_from"].isoformat() if turnover["period_from"] else None
    turnover["period_to"] = turnover["period_to"].isoformat() if turnover["period_to"] else None

    # --- непрерывность остатка: самая сильная проверка полноты выписки
    # Если остаток предыдущей строки ± сумма операции ≠ остаток текущей, между ними
    # пропущены операции. Это довод против доказательственной полноты выписки,
    # который иначе не увидеть: по датам «дырки» может не быть вовсе.
    balance_check = {"checked": False, "breaks": [], "coverage": None, "skipped_reason": None}
    with_balance = [o for o in ops if o["balance"] is not None]
    coverage = (len(with_balance) / len(ops)) if ops else 0
    balance_check["coverage"] = round(coverage, 3)
    dates_seq = [o["date"] for o in ops if o["date"]]
    descending = len(dates_seq) >= 3 and all(a >= b for a, b in zip(dates_seq, dates_seq[1:]))

    # Проверка осмысленна ТОЛЬКО когда остаток стоит почти в каждой строке и строки
    # идут по возрастанию даты. Иначе цепочка «предыдущий остаток ± эта операция»
    # рвётся на каждом пропуске, и профиль объявил бы неполной корректную выписку —
    # утверждение о доказательстве, которое юрист понесёт в суд.
    if len(with_balance) < 3:
        balance_check["skipped_reason"] = "остаток указан менее чем в трёх строках"
    elif coverage < 0.95:
        balance_check["skipped_reason"] = (
            f"остаток заполнен лишь в {coverage:.0%} строк — сверка цепочки невозможна "
            f"(вероятно, сальдо только на конец дня)")
        add("balance-not-checked", "info",
            f"Полнота выписки по остатку **не проверялась**: колонка остатка заполнена "
            f"в {coverage:.0%} строк из {len(ops)}. Это не значит, что выписка полна — "
            f"значит, машинно подтвердить полноту нечем. Запросите выписку с остатком "
            f"по каждой операции",
            [], {"coverage": balance_check["coverage"]})
    elif descending:
        balance_check["skipped_reason"] = "строки отсортированы по убыванию даты"
        add("balance-not-checked", "info",
            "Выписка отсортирована от новых операций к старым — сверка остатка по "
            "цепочке в таком порядке невозможна. Отсортируйте по возрастанию даты "
            "и повторите разбор", [], {})
    else:
        balance_check["checked"] = True
        for prev, cur in zip(with_balance, with_balance[1:]):
            expected = prev["balance"] + (cur["credit"] or 0) - (cur["debit"] or 0)
            if not _close(expected, cur["balance"], tolerance):
                gap = round(cur["balance"] - expected, 2)
                balance_check["breaks"].append({"rows": [prev["row"], cur["row"]],
                                                "gap": gap})
                if len(balance_check["breaks"]) <= 10:
                    add("balance-break", "error",
                        f"разрыв остатка между строками {prev['row']} и {cur['row']}: "
                        f"по предыдущему остатку и сумме операции должно быть "
                        f"{_money(expected)}, в выписке {_money(cur['balance'])} "
                        f"(расхождение {_money_signed(gap)}). Между строками пропущены "
                        f"операции — выписка неполная",
                        [_addr(sheet_name, cur["row"], cols.get("balance", 0))],
                        {"expected": round(expected, 2), "stated": cur["balance"],
                         "gap": gap})
        if len(balance_check["breaks"]) > 10:
            add("balance-break-many", "error",
                f"разрывов остатка всего {len(balance_check['breaks'])} — показаны первые 10. "
                f"Выписка систематически неполна", [], {"breaks": len(balance_check["breaks"])})

    # --- дробление: группа платежей одному лицу в узком окне
    # Группа собирается ЦЕЛИКОМ (жадно, пока платежи укладываются в окно от первого),
    # а не обрывается на достижении порога: иначе четвёртый платёж выпадает из группы
    # и в отчёте видно 3 платежа вместо 4 — сумма «дробления» занижается.
    splits = []
    by_party = {}
    for o in dated:
        if o["debit"] and o["party"]:
            by_party.setdefault(o["party"], []).append(o)
    for party, items in by_party.items():
        items.sort(key=lambda x: x["date"])
        i = 0
        while i < len(items):
            group = [items[i]]
            j = i + 1
            while j < len(items) and (items[j]["date"] - group[0]["date"]).days <= SPLIT_WINDOW_DAYS:
                group.append(items[j])
                j += 1
            if len(group) >= SPLIT_MIN_PAYMENTS:
                splits.append({"party": party,
                               "rows": [g["row"] for g in group],
                               "total": round(sum(g["debit"] for g in group), 2),
                               "from": group[0]["date"].isoformat(),
                               "to": group[-1]["date"].isoformat()})
                i = j
            else:
                i += 1
    split_rows = {r for s in splits for r in s["rows"]}
    for s in splits[:10]:
        same_day = s["from"] == s["to"]
        add("payment-splitting", "warn",
            f"«{s['party']}» — {_payments_word(len(s['rows']))}, "
            + (f"все {s['from']}" if same_day else f"за {s['from']}–{s['to']}")
            + f", на {_money(s['total'])} (строки {', '.join(map(str, s['rows']))}). "
              f"Похоже на дробление — гипотеза, требует проверки: у дробления бывают "
              f"законные причины (лимиты банка, разные счета-фактуры, график поставок)",
            [_addr(sheet_name, s["rows"][0], cols.get("date", 0))],
            {"count": len(s["rows"]), "total": s["total"]})

    # --- дубли: одна дата + одна сумма + один контрагент
    # Строки, уже объяснённые флагом дробления, повторно как «дубли» не подаём —
    # иначе одно явление даёт четыре флага и юрист перестаёт их читать.
    seen, duplicates, suppressed = {}, [], 0
    for o in ops:
        key = (o["date"], o["debit"], o["credit"], o["party"])
        if o["date"] is None or (o["debit"] is None and o["credit"] is None):
            continue
        if key in seen:
            pair = [seen[key], o["row"]]
            if all(r in split_rows for r in pair):
                suppressed += 1
                continue
            duplicates.append({"rows": pair, "amount": o["debit"] or o["credit"],
                               "party": o["party"]})
        else:
            seen[key] = o["row"]
    for d in duplicates[:10]:
        add("duplicate-operation", "warn",
            f"строки {d['rows'][0]} и {d['rows'][1]}: одинаковые дата, сумма "
            f"({_money(d['amount'])}) и контрагент — либо дубль в выписке, либо два "
            f"действительно одинаковых платежа. Проверьте перед использованием суммы",
            [_addr(sheet_name, r, cols.get("date", 0)) for r in d["rows"]])
    if len(duplicates) > 10:
        add("duplicate-operation-many", "warn",
            f"совпадающих операций всего {len(duplicates)} — показаны первые 10", [],
            {"duplicates": len(duplicates)})
    if suppressed:
        add("duplicate-in-splitting", "info",
            f"ещё {suppressed} совпадающих операций входят в группы, отмеченные как "
            f"возможное дробление, — отдельными флагами не дублируются", [],
            {"suppressed": suppressed})

    # --- крупнейшие получатели (куда ушли деньги)
    top = sorted(({"party": p,
                   "debit_total": round(sum(o["debit"] or 0 for o in items), 2),
                   "operations": len(items),
                   "rows": [o["row"] for o in items[:5]]}
                  for p, items in by_party.items()),
                 key=lambda x: x["debit_total"], reverse=True)[:TOP_PARTIES]

    # --- разрыв по датам: подозрение на неполноту, когда остатка в выписке нет
    date_gaps = []
    if dated and not balance_check["checked"]:
        dated.sort(key=lambda o: o["date"])
        for prev, cur in zip(dated, dated[1:]):
            gap_days = (cur["date"] - prev["date"]).days
            if gap_days >= 31:
                date_gaps.append({"rows": [prev["row"], cur["row"]], "days": gap_days,
                                  "from": prev["date"].isoformat(),
                                  "to": cur["date"].isoformat()})
        for g in date_gaps[:5]:
            add("date-gap", "info",
                f"между {g['from']} и {g['to']} нет операций ({g['days']} дн., строки "
                f"{g['rows'][0]}–{g['rows'][1]}). Само по себе не дефект, но проверьте, "
                f"полна ли выписка: колонки остатка в ней нет, поэтому машинно "
                f"подтвердить полноту невозможно",
                [_addr(sheet_name, g["rows"][1], cols.get("date", 0))],
                {"gap_days": g["days"]})

    # --- итоговая строка против суммы операций
    totals = {"stated_total": None, "stated_total_row": None, "delta": None,
              "compared_against": None}
    for i, row in enumerate(rows):
        joined = " ".join(_cell_str(c).lower() for c in row if not _cell_empty(c))
        if any(m in joined for m in ("итого", "всего", "оборот за период")):
            for key, ref in (("debit", turnover["debit_total"]),
                             ("credit", turnover["credit_total"])):
                if key in cols:
                    val = _num(row[cols[key]] if cols[key] < len(row) else None)
                    if val is None:
                        continue
                    totals.update({"stated_total": val,
                                   "stated_total_row": header_row1 + 1 + i,
                                   "compared_against": key,
                                   "delta": round(val - ref, 2)})
                    if not _close(val, ref, tolerance):
                        add("total-mismatch", "error",
                            f"итог по колонке «{headers[cols[key]]}» в таблице "
                            f"{_money(val)}, сумма операций {_money(ref)}, расхождение "
                            f"{_money_signed(val - ref)}",
                            [_addr(sheet_name, header_row1 + 1 + i, cols[key])],
                            {"stated": val, "computed": ref})
                    break
            break

    return {"columns": {k: _col_letter(v) for k, v in cols.items()},
            "turnover": turnover, "balance_check": balance_check,
            "duplicates": duplicates[:50], "splitting": splits[:50],
            "top_parties": top, "date_gaps": date_gaps[:50],
            "totals": totals, "findings": findings, "rows": []}


# --- профиль payments (платежи против периодов подозрительности) --------------
#
# Правовая рамка — гл. III.1 ФЗ-127, анкер сверен дословно 2026-07-30:
#   ст. 61.3 п. 2 — 1 месяц до принятия заявления (или после его принятия)
#   ст. 61.3 п. 3 — 6 месяцев до принятия заявления
#   ст. 61.2 п. 1 — 1 год до принятия заявления (или после)
#   ст. 61.2 п. 2 — 3 года до принятия заявления (или после)
# Анкер — ДАТА ПРИНЯТИЯ СУДОМ ЗАЯВЛЕНИЯ о признании должника банкротом; при
# нескольких заявлениях — дата принятия ПЕРВОГО (п. 7 ППВАС № 35 от 22.06.2012).
# Периоды ВЛОЖЕНЫ: платёж в пределах месяца попадает и в 6 месяцев, и в год, и в
# три года. Чем ближе к анкеру, тем легче состав: по ст. 61.3 п. 2 недобросовестность
# контрагента не доказывается (п. 11 ППВАС № 63), по ст. 61.2 п. 2 нужен полный
# состав из трёх элементов. Поэтому профиль называет САМЫЙ ЛЁГКИЙ применимый состав.

BORDER_DAYS = 5          # ± дней от границы периода — «на границе», решает один день

PAYMENT_MARKERS = {
    "date":       ("дата", "дата платежа", "дата операции", "дата сделки"),
    "amount":     ("сумма", "размер", "сумма платежа", "сумма сделки"),
    "party":      ("контрагент", "получатель", "кредитор", "наименование", "сторона"),
    "inn":        ("инн",),
    "purpose":    ("назначение", "основание", "содержание", "предмет"),
    "affiliated": ("аффилирован", "заинтересован", "связанн"),
}


def _shift_months(date: dt.date, months: int) -> dt.date:
    """Сдвиг на месяцы назад с коррекцией дня (31.03 − 1 мес → 28/29.02)."""
    year = date.year + (date.month - 1 - months) // 12
    month = (date.month - 1 - months) % 12 + 1
    day = date.day
    while day > 0:
        try:
            return dt.date(year, month, day)
        except ValueError:
            day -= 1
    return date


def _classify_payment(payment: dt.date, anchor: dt.date) -> dict:
    """Период подозрительности и самый лёгкий применимый состав."""
    bounds = {
        "1m": _shift_months(anchor, 1),
        "6m": _shift_months(anchor, 6),
        "1y": _shift_months(anchor, 12),
        "3y": _shift_months(anchor, 36),
    }
    if payment >= anchor:
        band, article = "после принятия заявления", "ст. 61.3 п. 2 / ст. 61.2"
    elif payment >= bounds["1m"]:
        band, article = "1 месяц до принятия заявления", "ст. 61.3 п. 2"
    elif payment >= bounds["6m"]:
        band, article = "6 месяцев до принятия заявления", "ст. 61.3 п. 3"
    elif payment >= bounds["1y"]:
        band, article = "1 год до принятия заявления", "ст. 61.2 п. 1"
    elif payment >= bounds["3y"]:
        band, article = "3 года до принятия заявления", "ст. 61.2 п. 2"
    else:
        band, article = "вне периодов гл. III.1", "только ст. 10 и 168 ГК (п. 4 ППВАС № 63)"

    near = []
    for key, bound in bounds.items():
        delta = abs((payment - bound).days)
        if delta <= BORDER_DAYS:
            near.append({"bound": key, "date": bound.isoformat(), "days": delta})
    return {"band": band, "article": article, "bounds": {k: v.isoformat() for k, v in bounds.items()},
            "near_border": near}


def profile_payments(sheet_name, headers, rows, header_row1, tolerance, anchor=None) -> dict:
    findings = []
    cols = _detect_by_markers(headers, PAYMENT_MARKERS)
    stats = {"rows_total": len(rows), "rows_used": 0, "rows_skipped": 0}

    def add(code, severity, message, cells=None, numbers=None):
        findings.append({"code": code, "severity": severity, "message": message,
                         "cells": cells or [], "numbers": numbers or {}})

    if "date" not in cols:
        for col in range(len(headers)):
            values = [r[col] for r in rows if col < len(r)]
            if sum(1 for v in values if _is_date_like(v)) >= max(2, len(values) // 3):
                cols["date"] = col
                break

    missing = [k for k in ("date", "amount") if k not in cols]
    if missing:
        add("columns-missing", "error",
            "Не распознаны обязательные колонки: " + ", ".join(missing)
            + ". Проверенные заголовки: " + (" | ".join(h for h in headers if h) or "(пусто)")
            + ". Периоды подозрительности НЕ рассчитаны — это отказ, а не «расхождений нет»")

    if anchor is None:
        add("anchor-missing", "error",
            "Не задана дата принятия судом заявления о признании должника банкротом — "
            "периоды подозрительности НЕ рассчитаны. Это не догадка инструмента: от этой "
            "даты зависит применимый состав, и ошибка на день меняет квалификацию. "
            "Передайте `--case-opened ГГГГ-ММ-ДД` либо `--case путь/к/case.yaml` "
            "(поле case.bankruptcy.case_opened_date). При нескольких заявлениях о "
            "банкротстве берите дату принятия ПЕРВОГО (п. 7 ППВАС № 35 от 22.06.2012)")

    payments, by_band = [], {}
    for i, row in enumerate(rows):
        row1 = header_row1 + 1 + i
        rec = {key: (row[idx] if idx < len(row) else None) for key, idx in cols.items()}
        date = _as_date(rec.get("date"))
        amount = _num(rec.get("amount"))
        if _is_total_row(row):
            continue
        if date is None or amount is None:
            if any(not _cell_empty(c) for c in row):
                stats["rows_skipped"] += 1
            continue
        stats["rows_used"] += 1
        item = {"row": row1, "date": date.isoformat(), "amount": amount,
                "party": _extract_party(rec), "inn": _cell_str(rec.get("inn")),
                "purpose": _cell_str(rec.get("purpose"))}
        if anchor:
            verdict = _classify_payment(date, anchor)
            item.update({"band": verdict["band"], "article": verdict["article"],
                         "near_border": verdict["near_border"]})
            bucket = by_band.setdefault(verdict["band"],
                                        {"count": 0, "total": 0.0, "article": verdict["article"],
                                         "rows": []})
            bucket["count"] += 1
            bucket["total"] = round(bucket["total"] + amount, 2)
            bucket["rows"].append(row1)
            for nb in verdict["near_border"]:
                add("near-period-border", "warn",
                    f"строка {row1}: платёж {date:%d.%m.%Y} на {_money(amount)} — "
                    f"в {nb['days']} дн. от границы периода «{nb['bound']}» "
                    f"({dt.date.fromisoformat(nb['date']):%d.%m.%Y}). От одного дня "
                    f"зависит применимый состав: проверьте дату по первичному документу, "
                    f"а не по реестру",
                    [_addr(sheet_name, row1, cols.get("date", 0))],
                    {"days_to_border": nb["days"], "bound": nb["bound"]})
        payments.append(item)

    if anchor and payments:
        outside = by_band.get("вне периодов гл. III.1")
        if outside:
            add("outside-periods", "info",
                f"{_payments_word(outside['count'])} на {_money(outside['total'])} — вне периодов "
                f"гл. III.1 (строки {', '.join(map(str, outside['rows'][:10]))}). По специальным "
                f"основаниям не оспариваются; остаётся ничтожность при злоупотреблении правом "
                f"(ст. 10 и 168 ГК, п. 4 ППВАС № 63) — другое основание и другая давность",
                [], {"count": outside["count"], "total": outside["total"]})
        easiest = by_band.get("1 месяц до принятия заявления") or \
            by_band.get("после принятия заявления")
        if easiest:
            add("easiest-composition", "info",
                f"{_payments_word(easiest['count'])} на {_money(easiest['total'])} попадают в "
                f"самый лёгкий состав — {easiest['article']}: предпочтение доказывается "
                f"фактом, недобросовестность контрагента доказывать не требуется "
                f"(п. 11 ППВАС № 63). Начинать разбор выгоднее с них",
                [], {"count": easiest["count"], "total": easiest["total"]})

    # аффилированность: один ИНН у разных наименований и наоборот
    if "inn" in cols:
        by_inn, by_name = {}, {}
        for p in payments:
            if p["inn"]:
                by_inn.setdefault(p["inn"], set()).add(p["party"])
            if p["party"]:
                by_name.setdefault(p["party"], set()).add(p["inn"])
        for inn, names in by_inn.items():
            names = {n for n in names if n}
            if len(names) > 1:
                add("inn-name-mismatch", "warn",
                    f"ИНН {inn} встречается с разными наименованиями: "
                    f"{', '.join(sorted(names))}. Либо переименование контрагента, либо "
                    f"ошибка в реестре — сверьте по ЕГРЮЛ, прежде чем строить довод",
                    [], {"inn": inn})

    if stats["rows_skipped"]:
        add("rows-skipped", "warn",
            f"Пропущено непустых строк: {stats['rows_skipped']} из {stats['rows_total']} "
            f"(не распознаны дата или сумма). Разобрано: {stats['rows_used']}. "
            f"Пропущенные платежи в периоды НЕ попали — проверьте формат этих строк",
            [], dict(stats))

    return {"columns": {k: _col_letter(v) for k, v in cols.items()},
            "stats": stats,
            "anchor": anchor.isoformat() if anchor else None,
            "bounds": ({k: v for k, v in _classify_payment(anchor, anchor)["bounds"].items()}
                       if anchor else None),
            "payments": payments, "by_band": by_band, "findings": findings, "rows": []}


# --- профиль registry (реестр требований кредиторов) --------------------------

REGISTRY_MARKERS = {
    "party":   ("кредитор", "наименование", "заявитель", "фио"),
    "inn":     ("инн",),
    "amount":  ("сумма", "размер требования", "включено", "основной долг", "требование"),
    "queue":   ("очередь", "очередность", "реестр"),
    "penalty": ("неустойка", "пени", "штраф", "финансовые санкции"),
    "date":    ("дата", "определение", "дата включения"),
    "voting":  ("голос", "голосующие", "% голосов"),
}

QUEUE_PATTERN = re.compile(r"(перв|втор|трет|четв|зареестр|текущ)", re.IGNORECASE)
# Реестры часто пишут очередь цифрой или римской цифрой — «3», «III», «3-я».
QUEUE_NUMERIC = {"1": "перв", "2": "втор", "3": "трет", "4": "четв",
                 "i": "перв", "ii": "втор", "iii": "трет", "iv": "четв"}


def _parse_queue(raw: str, row_text: str) -> str:
    """Очередь из ячейки. Числовые формы приводятся к словесным.

    По всей строке ищем только если своей ячейки нет: иначе кредитор
    «ООО "Первая грузовая"» уезжает в первую очередь (поймано проверкой).
    """
    value = (raw or "").strip().lower()
    if value:
        m = QUEUE_PATTERN.search(value)
        if m:
            return m.group(1).lower()
        token = re.sub(r"[^0-9ivх]", "", value.replace("х", "x"))
        if token in QUEUE_NUMERIC:
            return QUEUE_NUMERIC[token]
        return "не указана"
    m = QUEUE_PATTERN.search(row_text or "")
    return m.group(1).lower() if m else "не указана"


def profile_registry(sheet_name, headers, rows, header_row1, tolerance, our=None) -> dict:
    findings = []
    cols = _detect_by_markers(headers, REGISTRY_MARKERS)

    def add(code, severity, message, cells=None, numbers=None):
        findings.append({"code": code, "severity": severity, "message": message,
                         "cells": cells or [], "numbers": numbers or {}})

    if "amount" not in cols:
        add("columns-missing", "warn",
            "Колонка суммы требования не распознана — итоги по очередям и доля голосов "
            "не посчитаны. Заголовки: " + " | ".join(h for h in headers if h))

    creditors, queues = [], {}
    for i, row in enumerate(rows):
        row1 = header_row1 + 1 + i
        rec = {key: (row[idx] if idx < len(row) else None) for key, idx in cols.items()}
        amount = _num(rec.get("amount"))
        party = _cell_str(rec.get("party"))
        if amount is None and not party:
            continue
        joined = " ".join(_cell_str(c).lower() for c in row if not _cell_empty(c))
        if any(m in joined for m in ("итого", "всего")):
            continue
        queue = _parse_queue(_cell_str(rec.get("queue")), joined)
        item = {"row": row1, "party": party, "inn": _cell_str(rec.get("inn")),
                "amount": amount, "penalty": _num(rec.get("penalty")), "queue": queue}
        creditors.append(item)
        bucket = queues.setdefault(queue, {"count": 0, "total": 0.0, "penalty": 0.0,
                                           "rows": []})
        bucket["count"] += 1
        bucket["total"] = round(bucket["total"] + (amount or 0), 2)
        bucket["penalty"] = round(bucket["penalty"] + (item["penalty"] or 0), 2)
        bucket["rows"].append(row1)

    total_all = round(sum(c["amount"] or 0 for c in creditors), 2)

    # наше требование и вес голоса
    ours = None
    if our:
        needle = our.strip().lower()
        matched = [c for c in creditors
                   if needle in (c["party"] or "").lower() or needle == (c["inn"] or "")]
        if not matched:
            add("our-claim-missing", "error",
                f"Требование «{our}» в реестре НЕ НАЙДЕНО. Проверьте: включено ли оно "
                f"вообще, не искажено ли наименование, не пропущен ли срок предъявления "
                f"(за пределами срока требование становится зареестровым)",
                [], {"searched": our})
        else:
            our_total = round(sum(c["amount"] or 0 for c in matched), 2)
            ours = {"rows": [c["row"] for c in matched], "total": our_total,
                    "queue": matched[0]["queue"]}
            # Вес голоса считается по третьей очереди без финансовых санкций
            # (правила ст. 12 ФЗ-127). КЛЮЧЕВОЕ: вычитать неустойку можно только если
            # она входит в колонку суммы. Если колонка называется «Основной долг», а
            # неустойка стоит отдельно, вычитание занижает долю — на боевых числах это
            # 22,73 % вместо 25 %, то есть «блокирующего пакета нет» вместо «есть».
            third = queues.get("трет")
            amount_header = headers[cols["amount"]] if "amount" in cols else ""
            amount_is_principal = bool(re.search(
                r"основн|тело|долг(?!.*требован)", (amount_header or "").lower()))
            has_penalty_col = "penalty" in cols
            if third and third["total"]:
                if has_penalty_col and not amount_is_principal:
                    base = round(third["total"] - third["penalty"], 2)
                    our_base = round(our_total - sum(c["penalty"] or 0 for c in matched), 2)
                    basis = ("из суммы требований вычтены неустойки "
                             f"(колонка «{amount_header}» их включает)")
                elif has_penalty_col:
                    base, our_base = third["total"], our_total
                    basis = (f"колонка «{amount_header}» — основной долг, санкции учтены "
                             f"отдельной колонкой и в базу не входят")
                else:
                    base, our_base = third["total"], our_total
                    basis = ("⚠️ колонка неустоек НЕ найдена — если санкции включены в "
                             "сумму требований, доля ЗАВЫШЕНА; проверьте состав требований")
                if base <= 0:
                    add("voting-base-empty", "warn",
                        f"База голосов по третьей очереди равна {_money(base)} — "
                        f"голосующих требований нет либо очередь состоит только из "
                        f"санкций. Долю не считаю: результат был бы бессмысленным",
                        [], {"base": base})
                elif our_base < 0:
                    add("voting-base-negative", "warn",
                        f"У нашего требования неустойка ({_money(our_total - our_base)}) "
                        f"больше суммы в колонке «{amount_header}» ({_money(our_total)}). "
                        f"Похоже, колонки распознаны неверно — долю не считаю",
                        [], {"our_total": our_total, "our_base": our_base})
                else:
                    share = round(our_base / base * 100, 2)
                    ours.update({"voting_base": base, "our_voting_amount": our_base,
                                 "share_percent": share, "basis": basis})
                    add("voting-share", "info",
                        f"Доля нашего требования в третьей очереди: {share} % "
                        f"({_money(our_base)} из {_money(base)}). Как считано: {basis}. "
                        f"Пороги: более 25 % — блокировка решений, более 50 % — контроль "
                        f"собрания. Состав требований сверьте по определению о включении",
                        [], {"share_percent": share})
            else:
                add("third-queue-missing", "warn",
                    "Третья очередь в реестре не распознана — доля голоса не посчитана. "
                    "Проверьте колонку очереди: реестры пишут её и словом, и цифрой, и "
                    "римской цифрой", [], {})

    # дубли кредиторов и расхождения ИНН/наименования
    by_inn = {}
    for c in creditors:
        if c["inn"]:
            by_inn.setdefault(c["inn"], []).append(c)
    for inn, items in by_inn.items():
        if len(items) > 1:
            add("creditor-duplicate", "warn",
                f"ИНН {inn} встречается в {len(items)} строках "
                f"({', '.join(str(i['row']) for i in items)}) на общую сумму "
                f"{_money(sum(i['amount'] or 0 for i in items))}. Это либо несколько "
                f"требований одного кредитора (норма), либо задвоение — влияет и на "
                f"итог реестра, и на расчёт доли голосов",
                [], {"inn": inn, "rows": [i["row"] for i in items]})

    # итоговая строка против суммы
    stated_total, stated_row = None, None
    for i, row in enumerate(rows):
        joined = " ".join(_cell_str(c).lower() for c in row if not _cell_empty(c))
        if any(m in joined for m in ("итого", "всего")) and "amount" in cols:
            val = _num(row[cols["amount"]] if cols["amount"] < len(row) else None)
            if val is not None:
                stated_total, stated_row = val, header_row1 + 1 + i
                if not _close(val, total_all, tolerance):
                    add("total-mismatch", "error",
                        f"итог реестра в таблице {_money(val)}, сумма требований "
                        f"{_money(total_all)}, расхождение {_money_signed(val - total_all)}",
                        [_addr(sheet_name, stated_row, cols["amount"])],
                        {"stated": val, "computed": total_all})
                break

    return {"columns": {k: _col_letter(v) for k, v in cols.items()},
            "creditors_count": len(creditors), "total": total_all,
            "queues": queues, "ours": ours,
            "totals": {"stated_total": stated_total, "stated_total_row": stated_row},
            "findings": findings, "rows": []}


PROFILES = {"debt-calc": profile_debt_calc, "statement": profile_statement,
            "payments": profile_payments, "registry": profile_registry}


def selftest() -> int:
    """`--selftest` — проверка профиля на сгенерированной таблице, без внешних файлов.

    Нужна для боевого прогона в Cowork: подтверждает, что профиль ловит подставленные
    дефекты и НЕ шумит на корректном расчёте (ложная тревога дороже пропуска — юрист
    перестанет доверять флагам). Ожидаемые коды перечислены явно.
    """
    import tempfile
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        print("SELFTEST не выполнен: " + dependency_hint("openpyxl", exc))
        return 1

    tmp = tempfile.mkdtemp(prefix="analyze_table_selftest_")
    ok_all = True

    # 1) расчёт с четырьмя подставленными дефектами
    wb = Workbook(); ws = wb.active; ws.title = "Расчёт"
    ws.append(["Расчёт неустойки"]); ws.append([])
    ws.append(["Сумма долга", "с", "по", "Дней", "Ставка, %", "Пени"])
    base, rate = 1000000, 16.0
    good = base * rate / 100 * 31 / 365
    ws.append([base, dt.date(2026, 1, 15), dt.date(2026, 2, 14), 31, rate, round(good, 2)])
    ws.append([base, dt.date(2026, 2, 12), dt.date(2026, 3, 14), 31, rate, round(good, 2)])  # overlap
    ws.append([base, dt.date(2026, 3, 15), dt.date(2026, 4, 14), 30, rate, round(good, 2)])  # days
    ws.append([base, dt.date(2026, 4, 15), dt.date(2026, 5, 15), 31, rate, round(good + 5000, 2)])
    ws.append([None, None, None, None, "ИТОГО", round(good * 4 + 5000 + 10000, 2)])          # total
    bad_path = os.path.join(tmp, "bad.xlsx"); wb.save(bad_path)

    # 2) корректный расчёт (высокосный год → 366)
    wb2 = Workbook(); w2 = wb2.active; w2.title = "Проценты"
    w2.append(["Сумма долга", "с", "по", "Дней", "Ставка, %", "Проценты"])
    total = 0.0
    for start, end, days in ((dt.date(2024, 1, 10), dt.date(2024, 2, 9), 31),
                             (dt.date(2024, 2, 10), dt.date(2024, 3, 10), 30),
                             (dt.date(2024, 3, 11), dt.date(2024, 4, 10), 31)):
        val = round(500000 * 16.0 / 100 * days / 366, 2)
        total += val
        w2.append([500000, start, end, days, 16.0, val])
    w2.append([None, None, None, None, "ИТОГО", round(total, 2)])
    ok_path = os.path.join(tmp, "ok.xlsx"); wb2.save(ok_path)

    # 3) выписка: скрытые операции (разрыв остатка) + дробление + занижённый итог
    wb3 = Workbook(); w3 = wb3.active; w3.title = "Обороты"
    w3.append(["АО «Банк»"]); w3.append(["Выписка по счёту"]); w3.append([])
    w3.append(["Дата", "Контрагент", "Назначение", "Приход", "Списание", "Остаток"])
    bal = 5000000.0
    ledger = [(dt.date(2026, 2, 3), 'ООО "Поставщик"', None, 300000.0)]
    ledger += [(dt.date(2026, 2, 10 + (i // 3)), 'ООО "Ромашка"', None, 400000.0)
               for i in range(4)]                                    # дробление
    for date, party, credit, debit in ledger:
        bal += (credit or 0) - (debit or 0)
        w3.append([date, party, "оплата", credit, debit, round(bal, 2)])
    bal -= 2000000.0                                                 # скрытые операции
    w3.append([dt.date(2026, 3, 1), 'ООО "Прочий"', "возврат", None, 100000.0,
               round(bal - 100000, 2)])
    w3.append([None, None, "ИТОГО", None, 1000000.0, None])          # итог занижен
    st_path = os.path.join(tmp, "statement.xlsx"); wb3.save(st_path)

    for path, expect_codes, expect_errors in (
            (bad_path, {"period-overlap", "days-mismatch", "row-arithmetic", "total-mismatch"}, 5),
            (ok_path, set(), 0)):
        name, headers, rows, hdr = load_sheet(path)
        res = profile_debt_calc(name, headers, rows, hdr, DEFAULT_TOLERANCE)
        codes = {f["code"] for f in res["findings"] if f["severity"] == "error"}
        errors = sum(1 for f in res["findings"] if f["severity"] == "error")
        tag = Path(path).stem
        if codes != expect_codes or errors != expect_errors:
            ok_all = False
            print(f"SELFTEST [{tag}] ✗ ожидались коды {sorted(expect_codes) or '—'} "
                  f"и {expect_errors} ошибок; получено {sorted(codes) or '—'} / {errors}")
        else:
            print(f"SELFTEST [{tag}] ✓ коды {sorted(codes) or '—'}, ошибок {errors}"
                  + (f", формула «{res['formula']['label']}»" if res.get("formula") else ""))

    # 4) платежи против периодов подозрительности (анкер 15.06.2026)
    wb4 = Workbook(); w4 = wb4.active; w4.title = "Платежи"
    w4.append(["Дата", "Контрагент", "ИНН", "Назначение", "Сумма"])
    for date, party, inn, amount in (
            (dt.date(2026, 6, 1), 'ООО "Ромашка"', "7707083893", 1200000.0),   # 1 месяц
            (dt.date(2026, 5, 16), 'ООО "Лютик"', "7712345678", 500000.0),      # граница 1 дн.
            (dt.date(2026, 1, 10), 'ООО "Пион"', "7799999999", 2000000.0),      # 6 месяцев
            (dt.date(2024, 3, 3), 'ООО "Старый"', "7788888888", 5000000.0),     # 3 года
            (dt.date(2022, 1, 1), 'ООО "Древний"', "7766666666", 900000.0)):    # вне периодов
        w4.append([date, party, inn, "оплата", amount])
    pay_path = os.path.join(tmp, "payments.xlsx"); wb4.save(pay_path)

    # 5) реестр требований: наше требование, дубль ИНН, неустойки без голосов
    wb5 = Workbook(); w5 = wb5.active; w5.title = "Реестр"
    w5.append(["Кредитор", "ИНН", "Очередь", "Основной долг", "Неустойка"])
    for party, inn, queue, debt, penalty in (
            ('ООО "Наш клиент"', "7701111111", "третья", 50000000.0, 10000000.0),
            ('ПАО "Банк"', "7702222222", "третья", 120000000.0, 0.0),
            ('ФНС России', "7703333333", "третья", 30000000.0, 5000000.0),
            ('Работники', "", "вторая", 3000000.0, 0.0),
            ('ООО "Банк"', "7702222222", "третья", 20000000.0, 0.0)):
        w5.append([party, inn, queue, debt, penalty])
    reg_path = os.path.join(tmp, "registry.xlsx"); wb5.save(reg_path)

    # профиль statement
    name, headers, rows, hdr = load_sheet(st_path)
    res = profile_statement(name, headers, rows, hdr, DEFAULT_TOLERANCE)
    codes = {f["code"] for f in res["findings"] if f["severity"] == "error"}
    expect = {"balance-break", "total-mismatch"}
    split_groups = len(res["splitting"])
    split_size = max((len(s["rows"]) for s in res["splitting"]), default=0)
    if codes != expect or split_groups != 1 or split_size != 4:
        ok_all = False
        print(f"SELFTEST [statement] ✗ ожидались {sorted(expect)} и одна группа дробления "
              f"из 4 платежей; получено {sorted(codes)}, групп {split_groups}, "
              f"максимальная {split_size}")
    else:
        gap = res["balance_check"]["breaks"][0]["gap"]
        print(f"SELFTEST [statement] ✓ коды {sorted(codes)}, скрытых операций на "
              f"{_money(gap)}, дробление: 1 группа из 4 платежей")

    # профиль payments — сначала БЕЗ анкера: обязан отказаться считать
    name, headers, rows, hdr = load_sheet(pay_path)
    res = profile_payments(name, headers, rows, hdr, DEFAULT_TOLERANCE, anchor=None)
    if {f["code"] for f in res["findings"] if f["severity"] == "error"} != {"anchor-missing"}:
        ok_all = False
        print("SELFTEST [payments/без анкера] ✗ без даты принятия заявления профиль обязан "
              "вернуть ровно anchor-missing и не считать периоды")
    else:
        print("SELFTEST [payments/без анкера] ✓ отказался считать без даты принятия заявления")

    # с анкером: пять платежей должны разойтись по пяти разным периодам
    res = profile_payments(name, headers, rows, hdr, DEFAULT_TOLERANCE,
                           anchor=dt.date(2026, 6, 15))
    bands = res["by_band"]
    expected_bands = {"1 месяц до принятия заявления", "6 месяцев до принятия заявления",
                      "3 года до принятия заявления", "вне периодов гл. III.1"}
    border = [f for f in res["findings"] if f["code"] == "near-period-border"]
    if not expected_bands.issubset(set(bands)) or len(border) != 1:
        ok_all = False
        print(f"SELFTEST [payments] ✗ ожидались периоды {sorted(expected_bands)} и один флаг "
              f"границы; получено {sorted(bands)}, флагов границы {len(border)}")
    else:
        print(f"SELFTEST [payments] ✓ периодов {len(bands)}, границы: {len(border)}, "
              f"вне гл. III.1: {_money(bands['вне периодов гл. III.1']['total'])}")

    # профиль registry — доля голоса считается без неустоек
    name, headers, rows, hdr = load_sheet(reg_path)
    res = profile_registry(name, headers, rows, hdr, DEFAULT_TOLERANCE, our="Наш клиент")
    ours = res.get("ours") or {}
    dup = [f for f in res["findings"] if f["code"] == "creditor-duplicate"]
    # Арифметика ожидания — считана по НОРМЕ, а не по модели кода (прежнее ожидание
    # 19,51 % воспроизводило дефект: неустойка вычиталась из колонки, которая её не
    # включает). Колонка суммы здесь — «Основной долг», санкции стоят отдельно, значит
    # база = 50 + 120 + 30 + 20 = 220 млн, наше = 50 млн, доля = 50/220 = 22,73 %.
    if round(ours.get("share_percent") or 0, 2) != 22.73 or len(dup) != 1:
        ok_all = False
        print(f"SELFTEST [registry/основной долг] ✗ ожидалась доля 22.73 % и один дубль "
              f"ИНН; получено {ours.get('share_percent')} %, дублей {len(dup)}")
    else:
        print(f"SELFTEST [registry/основной долг] ✓ доля {ours['share_percent']} % — "
              f"неустойка НЕ вычтена (колонка её не включает), дубль ИНН найден")

    # Второй случай той же нормы: колонка «Сумма требования» неустойку ВКЛЮЧАЕТ —
    # тогда вычитать обязательно. Без этой фикстуры регресс в обратную сторону не виден.
    wb6 = Workbook(); w6 = wb6.active; w6.title = "Реестр"
    w6.append(["Кредитор", "ИНН", "Очередь", "Сумма требования", "в т.ч. неустойка"])
    for party, inn, queue, total_claim, penalty in (
            ('ООО "Наш клиент"', "7701111111", "3", 60000000.0, 10000000.0),
            ('ПАО "Банк"', "7702222222", "III", 140000000.0, 0.0),
            ('ФНС России', "7703333333", "третья", 35000000.0, 5000000.0)):
        w6.append([party, inn, queue, total_claim, penalty])
    incl_path = os.path.join(tmp, "registry_incl.xlsx"); wb6.save(incl_path)

    name, headers, rows, hdr = load_sheet(incl_path)
    res = profile_registry(name, headers, rows, hdr, DEFAULT_TOLERANCE, our="Наш клиент")
    ours2 = res.get("ours") or {}
    # база = (60+140+35) − (10+0+5) = 220 млн; наше = 60 − 10 = 50 млн → 22,73 %
    # плюс проверяем, что очередь распознана из «3» и «III», а не только из слова
    third = (res.get("queues") or {}).get("трет") or {}
    if round(ours2.get("share_percent") or 0, 2) != 22.73 or third.get("count") != 3:
        ok_all = False
        print(f"SELFTEST [registry/сумма требования] ✗ ожидалась доля 22.73 % и три "
              f"кредитора в третьей очереди; получено {ours2.get('share_percent')} %, "
              f"кредиторов {third.get('count')}")
    else:
        print(f"SELFTEST [registry/сумма требования] ✓ доля {ours2['share_percent']} % — "
              f"неустойка вычтена; очередь распознана из «3»/«III»/слова")

    # statement: оборот не должен включать строку «ИТОГО»
    name, headers, rows, hdr = load_sheet(st_path)
    res_st = profile_statement(name, headers, rows, hdr, DEFAULT_TOLERANCE)
    debit = res_st["turnover"]["debit_total"]
    if abs(debit - 2000000.0) > 0.01:
        ok_all = False
        print(f"SELFTEST [statement/оборот] ✗ списание должно быть 2 000 000 (5 операций), "
              f"получено {debit} — вероятно, строка «ИТОГО» посчитана операцией")
    else:
        print("SELFTEST [statement/оборот] ✓ строка «ИТОГО» операцией не считается")

    print("SELFTEST:", "ВСЁ ЗЕЛЁНОЕ" if ok_all else "ЕСТЬ ОТКАЗЫ")
    return 0 if ok_all else 1


def main():
    # Как в extract_text.py: на Windows stdout по умолчанию cp1251 и падает на
    # кириллице/типографике в JSON. Кодировку задаём явно, ensure_ascii=False
    # оставляем — JSON должен быть читаемым, а не в \uXXXX.
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

    argv = sys.argv[1:]
    if "--selftest" in argv:
        return selftest()
    if not argv or argv[0].startswith("--"):
        print(json.dumps({"error": "укажите файл: analyze_table.py <файл.xlsx> "
                                   "--profile debt-calc  (или --selftest)"},
                         ensure_ascii=False))
        return 2
    path = argv[0]

    def opt(name, default=None):
        return argv[argv.index(name) + 1] if name in argv and argv.index(name) + 1 < len(argv) \
            else default

    profile = opt("--profile", "debt-calc")
    sheet = opt("--sheet")
    global _REL_TOLERANCE_ACTIVE
    if "--tolerance" in argv:
        _REL_TOLERANCE_ACTIVE = None        # явный допуск отключает относительный
    try:
        tolerance = float(str(opt("--tolerance", DEFAULT_TOLERANCE)).replace(",", "."))
    except ValueError:
        print(json.dumps({"error": "--tolerance: ожидается число, например 1 или 0.5"},
                         ensure_ascii=False))
        return 2

    if profile not in PROFILES:
        print(json.dumps({"error": f"неизвестный профиль: {profile}",
                          "available": sorted(PROFILES)}, ensure_ascii=False))
        return 2
    if not os.path.exists(path):
        print(json.dumps({"error": f"файл не найден: {path}"}, ensure_ascii=False))
        return 1

    try:
        name, headers, rows, header_row1 = load_sheet(path, sheet)
    except Exception as exc:
        print(json.dumps({"error": f"не удалось прочитать таблицу: {exc}"},
                         ensure_ascii=False))
        return 1

    if not rows:
        print(json.dumps({"profile": profile, "sheet": name, "error": "лист пуст"},
                         ensure_ascii=False))
        return 0

    extra = {}
    anchor_problem = None
    if profile == "payments":
        raw_opened = opt("--case-opened")
        anchor = _as_date(raw_opened)
        if raw_opened and anchor is None:
            anchor_problem = (f"--case-opened «{raw_opened}» не распознан как дата; "
                              f"формат ГГГГ-ММ-ДД или ДД.ММ.ГГГГ")
        case_path = opt("--case")
        anchor_source = "--case-opened" if anchor else None
        if anchor is None and case_path:
            anchor, anchor_source, anchor_problem = _anchor_from_case(case_path)
        extra["anchor"] = anchor
        result_anchor_source = anchor_source
    elif profile == "registry":
        extra["our"] = opt("--our")
        result_anchor_source = None
    else:
        result_anchor_source = None

    result = PROFILES[profile](name, headers, rows, header_row1, tolerance, **extra)
    if result_anchor_source:
        result["anchor_source"] = result_anchor_source
    if anchor_problem:
        result.setdefault("findings", []).insert(0, {
            "code": "anchor-source-failed", "severity": "error",
            "message": f"Дата принятия заявления передана, но не прочитана: {anchor_problem}. "
                       f"Это НЕ «дата не задана» — исправьте ввод и повторите",
            "cells": [], "numbers": {}})

    # Формулы без сохранённых значений — проверяем и здесь (юрист мог запустить ногу
    # по файлу, не проходившему через приём).
    if not str(path).lower().endswith(".csv"):
        stale = check_stale_formulas(path, sheet)
        result["stale_formulas"] = stale
        if stale.get("stale"):
            where = ", ".join(f"«{e['sheet']}» стр. {e['row']}"
                              for e in stale["examples"][:3])
            result.setdefault("findings", []).insert(0, {
                "code": "stale-formulas", "severity": "error",
                "message": f"В файле {stale['stale']} ячеек с формулами БЕЗ сохранённых "
                           f"значений ({where}). Прочитанным суммам верить нельзя: в Excel "
                           f"они видны, в файле их нет. Откройте файл, пересохраните и "
                           f"повторите разбор. Всё посчитанное ниже — недостоверно",
                "cells": [], "numbers": {"stale": stale["stale"]}})

    # Разбирается ОДИН лист; если их несколько — сказать прямо.
    sheets_seen = getattr(load_sheet, "last_sheets", None)
    if sheets_seen and len(sheets_seen) > 1 and not sheet:
        result.setdefault("findings", []).append({
            "code": "single-sheet-analyzed", "severity": "warn",
            "message": f"В файле {len(sheets_seen)} листов ({', '.join(sheets_seen)}), "
                       f"разобран только «{name}». Итог относится к нему одному — "
                       f"для остальных запустите с --sheet",
            "cells": [], "numbers": {"sheets": len(sheets_seen)}})
    result.update({"profile": profile, "sheet": name, "file": Path(path).name,
                   "header_row": header_row1,
                   "errors": sum(1 for f in result["findings"] if f["severity"] == "error")})

    out_dir = opt("--out-dir")
    payload = json.dumps(result, ensure_ascii=False, indent=2)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
        dst = os.path.join(out_dir, Path(path).stem + f".{profile}.json")
        with open(dst, "w", encoding="utf-8") as f:
            f.write(payload)
        print(json.dumps({"saved": dst, "errors": result["errors"],
                          "findings": len(result["findings"])}, ensure_ascii=False))
    else:
        print(payload)
    return 0


if __name__ == "__main__":
    sys.exit(main())
